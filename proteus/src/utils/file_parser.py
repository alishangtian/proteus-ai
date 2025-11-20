import os
import logging
import json
from typing import Optional, Dict, Any
from src.api.llm_api import call_multimodal_llm_api
from src.utils.redis_cache import get_redis_connection
from src.utils.logger import setup_logger
from dotenv import load_dotenv
from PyPDF2 import PdfReader  # 导入 PyPDF2
from docx import Document  # 导入 python-docx
from openpyxl import load_workbook
import xlrd

load_dotenv()
language = os.getenv("LANGUAGE", "中文")

log_file_path = os.getenv("log_file_path", "logs/workflow_engine.log")
setup_logger(log_file_path)
logger = logging.getLogger(__name__)


async def parse_file(file_path: str, file_type: str, file_id: str) -> Optional[str]:
    """
    根据文件类型解析文件内容。
    Args:
        file_path: 文件的完整路径。
        file_type: 文件的MIME类型。
        file_id: 文件的唯一ID。
    Returns:
        解析后的文件内容字符串，如果无法解析则返回None。
    """
    logger.info(f"开始解析文件: {file_path}, 类型: {file_type}")
    parsed_content = None

    # 获取文件扩展名
    file_extension = os.path.splitext(file_path)[1].lower()

    # 优先根据文件后缀判断文本类型文件
    if file_extension in [".txt", ".md", ".csv"]:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
            parsed_content = text
            logger.info(f"文本文件 '{file_path}' (后缀: {file_extension}) 解析成功")
        except Exception as e:
            logger.error(f"文本文件 '{file_path}' 读取失败: {str(e)}", exc_info=True)
            parsed_content = f"文本文件读取失败: {str(e)}"
    elif file_type.startswith("image/"):
        try:
            messages = [
                {
                    "role": "user",
                    "content": f"请详细使用{language}描述这张图片的内容",
                },
                {
                    "role": "user",
                    "content": [{"type": "image_url", "image_url": {"url": file_path}}],
                },
            ]
            resp, _ = await call_multimodal_llm_api(
                messages=messages, request_id=file_id
            )
            parsed_content = resp
            logger.info(f"图片文件 '{file_path}' 解析成功")
        except Exception as e:
            logger.error(f"图片文件 '{file_path}' 解析失败: {str(e)}", exc_info=True)
            parsed_content = f"图片解析失败: {str(e)}"
    elif file_type.startswith("text/") or file_type in [
        "application/json",
        "application/xml",
        "application/javascript",
        "text/markdown",
        "text/plain",
        "text/html",
        "application/x-python-code",  # .py files
    ]:
        try:
            text = ""
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
            parsed_content = text
            logger.info(f"文本文件 '{file_path}' 解析成功")
        except Exception as e:
            logger.error(f"文本文件 '{file_path}' 读取失败: {str(e)}", exc_info=True)
            parsed_content = f"文本文件读取失败: {str(e)}"
    elif file_extension in [".xlsx", ".xls"] or file_type in [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]:
        # 增加 Excel 解析：优先解析第一个 sheet，输出为 Markdown 表格
        try:
            rows = []
            if file_extension == ".xlsx" and load_workbook is not None:
                wb = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet = wb.worksheets[0]
                rows = list(sheet.iter_rows(values_only=True))
            elif file_extension == ".xls" and xlrd is not None:
                book = xlrd.open_workbook(file_path)
                sheet = book.sheet_by_index(0)
                rows = [sheet.row_values(i) for i in range(sheet.nrows)]
            else:
                # 如果缺少依赖，抛出错误以便被捕获并记录
                raise ImportError(
                    "缺少解析 Excel 的依赖（openpyxl/xlrd）。请安装 openpyxl（用于 .xlsx）或 xlrd（用于 .xls）"
                )

            # 规范化表格行：转换为字符串，补齐列宽
            max_cols = max((len(r) for r in rows), default=0)
            norm_rows = []
            for r in rows:
                row = []
                for i in range(max_cols):
                    val = r[i] if i < len(r) else ""
                    if val is None:
                        row.append("")
                    else:
                        # 去除换行，转换为字符串
                        cell_text = str(val).replace("\n", " ").strip()
                        row.append(cell_text)
                norm_rows.append(row)

            # 构建 Markdown 表格
            if not norm_rows:
                parsed_content = ""
            else:
                header = norm_rows[0]
                # 如果 header 全为空，则生成通用列名
                if all((cell.strip() == "" for cell in header)):
                    header = [f"列{i+1}" for i in range(max_cols)]
                sep = ["---"] * max_cols
                lines = []
                lines.append("| " + " | ".join(header) + " |")
                lines.append("| " + " | ".join(sep) + " |")
                for row in norm_rows[1:]:
                    lines.append("| " + " | ".join(row) + " |")
                parsed_content = "\n".join(lines)
            logger.info(f"Excel 文件 '{file_path}' 解析成功，已转换为 Markdown 表格")
        except Exception as e:
            logger.error(f"Excel 文件 '{file_path}' 解析失败: {str(e)}", exc_info=True)
            parsed_content = f"Excel 文件解析失败: {str(e)}"
    elif file_type == "application/pdf":
        try:
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            parsed_content = text
            logger.info(f"PDF 文件 '{file_path}' 解析成功")
        except Exception as e:
            logger.error(f"PDF 文件 '{file_path}' 解析失败: {str(e)}", exc_info=True)
            parsed_content = f"PDF 文件解析失败: {str(e)}"
    elif (
        file_type
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        try:
            document = Document(file_path)
            text = "\n".join([para.text for para in document.paragraphs])
            parsed_content = text
            logger.info(f"DOCX 文件 '{file_path}' 解析成功")
        except Exception as e:
            logger.error(f"DOCX 文件 '{file_path}' 解析失败: {str(e)}", exc_info=True)
            parsed_content = f"DOCX 文件解析失败: {str(e)}"
    else:
        logger.info(f"文件 '{file_path}' 类型 '{file_type}' 不支持解析，只进行上传。")
        parsed_content = None  # 不支持解析，返回 None

    return parsed_content
